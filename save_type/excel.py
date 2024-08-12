import pandas as pd
import os
from openpyxl import load_workbook

def save_to_excel(image_imports, output_file, type):
    # Ensure the output directory exists
    output_dir = os.path.dirname(output_file)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Prepare data for DataFrame
    data = []
    for file, imports in image_imports.items():
        for imp in imports:
            if len(imp) == 2:  # For MUI icons
                data.append({'File': file, 'type': type, 'Name': imp[1]})
            else:  # For images
                data.append({'File': file, 'type': type, 'Name': imp[0], 'Path': f'/public/{imp[1]}'})

    # Create DataFrame
    df = pd.DataFrame(data)

    # Save DataFrame to Excel file
    df.to_excel(output_file, index=False, engine='openpyxl')

    # Load the Excel file with openpyxl
    wb = load_workbook(output_file)
    ws = wb.active

    # Merge cells for the 'File' column
    file_column = 'A'  # Assuming 'File' is in column A
    start_row = 2  # Assuming there is a header row

    # Loop through rows and merge cells with the same file path
    current_file = None
    start_merge_row = start_row
    for row in range(start_row, ws.max_row + 1):
        cell_value = ws[f'{file_column}{row}'].value
        if cell_value == current_file:
            # Merge cells in the 'File' column
            ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=row - 1, end_column=1)
        else:
            current_file = cell_value
            start_merge_row = row

    # Save the updated Excel file
    wb.save(output_file)

    print(f"Data with merged cells has been saved to {output_file}")
